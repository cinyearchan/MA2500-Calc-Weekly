import pandas as pd
import baostock as bs
import datetime as dt
import openpyxl as xl

lj1 = './上证指数20年汇总.xlsx'
lj2 = './上证指数10年分布统计.xlsx'

end = dt.datetime.utcnow().strftime('%Y-%m-%d')  # 结束时间
start = '2001-01-01'  # 开始时间

lg = bs.login()  # 登录系统
print('login respond error_code: ' + lg.error_code)  # lg.error_code 错误代码，当为 0 时表示成功，当为 非0 时表示失败
print('login respond error_msg: ' + lg.error_msg)  # lg.error_msg 错误信息，对错误的详细解释

sh_rs = bs.query_history_k_data_plus(
    'sh.000001',
    'date,code,close',
    start_date=start,
    end_date=end,
    frequency='d',
    adjustflag='3'
)  # peTTM滚动市盈率 pbMRQ市净率 frequency=d日K线 adjustflag=3 不复权

sh_list = []

while (sh_rs.error_code == '0') & sh_rs.next():
    sh_list.append(sh_rs.get_row_data())

sh_result = pd.DataFrame(sh_list, columns=sh_rs.fields)
sh_result.to_excel(lj1, index=False)
lg = bs.logout()  # 登录系统

sh_series = sh_result['close'].astype(float)  # astype 转换数据类型
wb1 = xl.load_workbook(lj1)
sheet1 = wb1['Sheet1']
sheet1.cell(1, 4).value = 'MA2500均线'
sheet1.cell(1, 5).value = '极度低估'
sheet1.cell(1, 6).value = '比较便宜'
sheet1.cell(1, 7).value = '估值合理'
sheet1.cell(1, 8).value = '轻度泡沫'
sheet1.cell(1, 9).value = '高度泡沫'
sheet1.cell(1, 10).value = '玩命'
sheet1.freeze_panes = 'D2'  # 冻结单元格
sheet1.column_dimensions['A'].width = 12  # 设置列宽
sheet1.column_dimensions['B'].width = 12
sheet1.column_dimensions['C'].width = 12
sheet1.column_dimensions['D'].width = 12
a,b,c,d,e,f = 0,0,0,0,0,0
g = sheet1.max_row

for i in range(g-2499, g + 1):
    sheet1.cell(i, 4).value = sh_series.iloc[i-2499-2:i-2+1].mean()  # mean 均值 iloc 从0开始计数，不包含表头 列表包含开头
    if float(sheet1.cell(i, 3).value) < sheet1.cell(i, 4).value / 1.2:
        sheet1.cell(i, 5).value = 'Y'
        a = a + 1
    elif sheet1.cell(i, 4).value / 1.2 <= float(sheet1.cell(i, 3).value) < sheet1.cell(i, 4).value:
        sheet1.cell(i, 6).value = 'Y'
        b = b + 1
    elif sheet1.cell(i, 4).value <= float(sheet1.cell(i, 3).value) < sheet1.cell(i, 4).value * 1.2:
        sheet1.cell(i, 7).value = 'Y'
        c = c + 1
    elif sheet1.cell(i, 4).value * 1.2 <= float(sheet1.cell(i, 3).value) < sheet1.cell(i, 4).value * 1.4:
        sheet1.cell(i, 8).value = 'Y'
        d = d + 1
    elif sheet1.cell(i, 4).value * 1.4 <= float(sheet1.cell(i, 3).value) < sheet1.cell(i, 4).value * 1.6:
        sheet1.cell(i, 9).value = 'Y'
        e = e + 1
    elif sheet1.cell(i, 4).value * 1.6 <= float(sheet1.cell(i, 3).value):
        sheet1.cell(i, 10).value = 'Y'
        f = f + 1

sheet1.cell(g + 1, 5).value = a
sheet1.cell(g + 2, 5).value = a / 2500
sheet1.cell(g + 1, 6).value = b
sheet1.cell(g + 2, 6).value = b / 2500
sheet1.cell(g + 1, 7).value = c
sheet1.cell(g + 2, 7).value = c / 2500
sheet1.cell(g + 1, 8).value = d
sheet1.cell(g + 2, 8).value = d / 2500
sheet1.cell(g + 1, 9).value = e
sheet1.cell(g + 2, 9).value = e / 2500
sheet1.cell(g + 1, 10).value = f
sheet1.cell(g + 2, 10).value = f / 2500
wb1.save(lj2)
